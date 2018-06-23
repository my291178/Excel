
import xlrd, xlwt
import openpyxl
from xlwt import *

from openpyxl import Workbook, load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image


thick_border = Border(left=Side(style='thick'),
                      right=Side(style='thick'),
                      top=Side(style='thick'),
                      bottom=Side(style='thick'))


# wb = load_workbook('report.xlsx')
wb = Workbook()
ws = wb.active
ws.title = 'test'

ws['A1'] = "ОТЧЕТ"
ws.merge_cells('A1:D1')

# ws.add_image()
# ws.min_column

ws['A1'].border = thick_border
ws['B1'].border = thick_border
ws['C1'].border = thick_border
ws['D1'].border = thick_border

ws['A2'] = "ID"
ws['A2'].border = thick_border

ws['B2'] = "Name"
ws['B2'].border = thick_border

ws['C2'] = "Date"
ws['C2'].border = thick_border

ws['D2'] = "Price"
ws['D2'].border = thick_border



# w = Workbook()
# ws = wb.add_sheet('Image')
# ws.add_image('IMG_1003.jpg', anchor=None)

# Also works if you already have the image bitmap data in memory...
# with open ("IMG_1003.jpg", "r") as bmpfile:
#     bmpdata = bmpfile.read()
#     ws.insert_bitmap_data(bmpdata, 10, 2)

ws['A4'] = 'You should see three logos below'

# create an image
img = Image('boombob.bmp')

# add to worksheet and anchor next to cells
ws.add_image(img, 'A4')

wb.save('report.xlsx')
# w.save('image.xls')